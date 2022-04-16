import os, re, sys, time

from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter.ttk import *

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import ActionChains

from selenium.common.exceptions import InvalidSessionIdException, NoSuchElementException

APP_WINDOW = None
TXT_SESSIONID = None
STATUS = None
WORK_PAUSED = None
FILE_NAME = None
PROGRESS = None
PROGRESSBAR = None
GIS = None
GLOBAL_RETRIES_COUNT = 0

# ------------------------------------------------------------ #
# ------------------------------------------------------------ #
# ------------------------------------------------------------ #
def app_resize(event):
    with open("app.ini", "w") as conf:
        conf.write(APP_WINDOW.geometry())

# ------------------------------------------------------------ #
def print_status(msg, win_update=True):
    txt = msg
    print(txt)
    STATUS.set(txt)
    if win_update: APP_WINDOW.update()

# ------------------------------------------------------------ #
def print_progress(total=0, done=0, done_batch=0, duration=0, message=None, win_update=True):
    if message is not None:
        PROGRESS.set(message)
        print(message)
        APP_WINDOW.update()
        return
    msg = f"Всего в файле строк: {total}, из них обработано: {done}"
    if done_batch > 0:
        msg = msg + f"\nв т.ч. {done_batch}" + \
            (f" за {round(duration, 2)} сек." if duration > 0 else "") + \
            (f" ~{int(3600 / (duration / done_batch))} строк в час" if duration > 0 else "")
    print(msg)
    PROGRESS.set(msg)
    if msg != "":
        PROGRESSBAR['value'] = int(done / total * 100)
        PROGRESSBAR.grid()
    else:
        PROGRESSBAR.grid_remove()
    if win_update: APP_WINDOW.update()

# ------------------------------------------------------------ #
def main():
    global APP_WINDOW, TXT_SESSIONID, STATUS, FILE_NAME, PROGRESS, PROGRESSBAR

    APP_WINDOW = Tk()
    FILE_NAME = StringVar()
    PROGRESS = StringVar()
    STATUS = StringVar()

    APP_WINDOW.title("ГИС ЖКХ - проверка ответов на запросы по льготам")
    if os.path.isfile("app.ini"):
        with open("app.ini", "r") as conf:
            APP_WINDOW.geometry(conf.read())
    else:
        APP_WINDOW.geometry("600x300")
    APP_WINDOW.resizable(1, 1)
    APP_WINDOW.call("wm", "attributes", ".", "-topmost", True)
    APP_WINDOW.bind("<Configure>", app_resize)

    lbl_status = Label(APP_WINDOW, textvariable=STATUS)
    lbl_status.grid(row=1, column=1, columnspan=3)

    btn_1 = Button(APP_WINDOW, text="Запустить Chrome", command=GISinit, width=25)
    btn_1.grid(row=2, column=1, sticky="we")

    btn_2 = Button(APP_WINDOW, text="Выбрать файл: ", command=get_file_to_proceed)
    btn_2.grid(row=3, column=1, sticky="we")

    txt_filename = Label(APP_WINDOW, textvariable=FILE_NAME)
    txt_filename.grid(row=3, column=2, columnspan=3, sticky="w")

    lbl_progress = Label(APP_WINDOW, textvariable=PROGRESS)
    lbl_progress.grid(row=4, column=2, columnspan=3, sticky="we")

    PROGRESSBAR = Progressbar(APP_WINDOW, orient='horizontal', mode='determinate')
    PROGRESSBAR.grid(row=5, column=2, columnspan=2, sticky="we")
    PROGRESSBAR['value'] = 0
    PROGRESSBAR.grid_remove()

    TXT_SESSIONID = Entry(APP_WINDOW, justify="center")
    TXT_SESSIONID.grid(row=6, column=1, sticky="we")

    lbl_cookie = Label(APP_WINDOW, text="<<< sessionid из основного браузера")
    lbl_cookie.grid(row=6, column=2, columnspan=3, sticky="w")

    btn_3 = Button(APP_WINDOW, text="Начать работу", command=GISgo, width=15)
    btn_3.grid(row=7, column=1, sticky="we")

    btn_4 = Button(APP_WINDOW, text="Пауза", command=GISpause, width=15)
    btn_4.grid(row=8, column=1, sticky="we")

    btn_5 = Button(APP_WINDOW, text="Выход", command=GISquit)
    btn_5.grid(row=9, column=3, sticky="we")

    lbl_about = Label(APP_WINDOW, text="Бот для проверки ответов на запросы по льготам ГИС ЖКХ :: версия 2.05")
    lbl_about.grid(row=10, column=1, columnspan=3, sticky="w")

    Grid.columnconfigure(APP_WINDOW, 0, minsize=25)
    Grid.columnconfigure(APP_WINDOW, 1, minsize=150)
    Grid.columnconfigure(APP_WINDOW, 2, minsize=200)
    Grid.columnconfigure(APP_WINDOW, 3, minsize=150)
    Grid.columnconfigure(APP_WINDOW, 4, minsize=25, weight=1)
    APP_WINDOW.mainloop()

##------------------------------------------------------------##
def get_file_to_proceed():
    filename = askopenfilename(parent=APP_WINDOW,
                               title="Выберите файл для обработки",
                               filetypes=[("Файлы Excel", "*.xlsx")], multiple=False)
    if filename == "":
        FILE_NAME.set("Файл не выбран")
        print_status("Файл не выбран")
        return
    s = filename
    r = s.rfind("/")
    if r > 0: s = s[:r] + "\n" + s[r + 1:]
    FILE_NAME.set(s)
    PROGRESS = check_file_to_proceed(filename)
    print_progress(total=PROGRESS[0], done=PROGRESS[1], message=PROGRESS[2], win_update=False)
    print_status("Файл выбран", False)
    APP_WINDOW.update()

# ------------------------------------------------------------ #
def check_file_to_proceed(wbName):
    global SEARCH_TYPE
    try:
        wb = load_workbook(filename=wbName)
        ws = wb.worksheets[0]
        rmax = ws.max_row
    except:
        return [None, None, "Проблема с файлом"]

## пустой файл
    if rmax == 1: SEARCH_TYPE = None; return [None, None, "Файл пустой"]

## количество обработанных строк
    for row in range(2, rmax + 1):
        if ws.cell(row=row, column=2).value is None: row = row - 1; break

    if row == rmax: SEARCH_TYPE = None; return [None, None, "Файл полностью обработан"]
    else: return [rmax - 1, row - 1, None]

# ------------------------------------------------------------ #
def GISinit(message=None):
    global GIS

    if message is not None:
        print_status(message)

    try: GIS.close()
    except: pass

    try: GIS.quit()
    except: pass

    if len(TXT_SESSIONID.get()) != 34:
        print_status("Сначала нужно ввести sessionid из основного браузера")
        return

    chrome_options = Options()

    if not RUN_FROM_IDLE:
        chrome_options.add_argument("--headless")
##    chrome_options.add_argument("user-data-dir=" + os.getenv("TEMP") + "\\gis")

    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    GIS = webdriver.Chrome(options=chrome_options)
    GIS.set_page_load_timeout(180)
    print_status("Начало работы с сайтом")
    GIS.get("https://dom.gosuslugi.ru/404")
    GIS.add_cookie({'name': 'sessionid', 'value': TXT_SESSIONID.get(), 'domain': '.dom.gosuslugi.ru'})
    print_status("Загрузка первой страницы (может быть довольного долго)")
    GIS.get("https://my.dom.gosuslugi.ru/organization-cabinet/#!/debts/received-requests")
    print_status("ГИС успешно загружена")

def log(s):
    with open('log.txt', 'a') as f:
        f.write(f'{s}\n')

# ------------------------------------------------------------ #
def GISquit():
    global GIS

    print_status("Программа закрывается, подождите...")

    try: GIS.close()
    except: pass

    try: GIS.quit()
    except: pass

    sys.exit()

# ------------------------------------------------------------ #
def GISpause():
    global WORK_PAUSED
    WORK_PAUSED = True

    print_status("...пауза...")

# ------------------------------------------------------------ #
def GISgo():
    global WORK_PAUSED
    WORK_PAUSED = False

    if GIS is None: print_status("Сайт ещё не загружен"); return
    print_status("...идёт обработка файла...")

    t0 = time.monotonic()

## начало обработки файла
    wbName = FILE_NAME.get().replace("\n", "/")
    try: wb = load_workbook(filename=wbName); ws=wb.worksheets[0]
    except: print_status("Проблема с файлом: невозможно открыть и т.д."); return

    rows_total = ws.max_row - 1
    if rows_total < 1: print_status("Пустой файл"); return

## пропустить обработанные строки
    for row in range(2, rows_total + 1 + 1):
        if ws.cell(row=row, column=2).value is None: row = row - 1; break
    rows_done = row - 1
    if rows_done == rows_total:  print_status("Файл уже полностью обработан"); return

## добавить заголовок
    ws.cell(row=1, column=1).value = "Номер запроса"
    ws.cell(row=1, column=2).value = "Статус запроса/ответа"

## в фильтре поиска очистить поле Исполнитель
    try:
        filter_field = GIS.find_element(By.XPATH, ".//label[contains(text(),' Исполнитель ')]/..")
        while True:
            filters = filter_field.find_elements(By.CSS_SELECTOR, "ul > li > a")
            if len(filters) == 0: break
            filters[0].click()
    except:
        pass

## основной цикл обработки
    row = row + 1
    while True:
        if WORK_PAUSED: return
        cc = ws.cell(row=row, column=1).value
        if cc is None:
            print_status(f"Файл полностью обработан за {round(time.monotonic() - t0, 2)} сек.")
            print_progress(message="")
            break
        t = GISgo1_plusplus(str(cc).strip())
## разбор статусов запросов-ответов
        if t == "FATAL":
            print_status(f"Фатальная ошибка. Скорее всего, нужна новая сессия в ГИС ЖКХ.")
            return
        elif t == "ERROR":
            print_status(f"Ошибка. Пробуем перезапустить бота, попытка №{GLOBAL_RETRIES_COUNT}")
            row = row - 1
            GISinit("Перезапуск Chrome после фатальной ошибки")
            continue

        tt = t.split(";")
        print_status(", ".join(tt[:4]))
        print_progress(total=rows_total,
                       done=row - 1,
                       done_batch=row - 1 - rows_done,
                       duration=time.monotonic() - t0)
        ws.cell(row=row, column=2).value = tt[1]
        if tt[1] == "Ответ не отправлен":
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="00FFCC99")
        ws.cell(row=row, column=3).value = tt[2]
        ws.cell(row=row, column=4).value = tt[3]
        if tt[3] == "Файлов нет":
            ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor="00FF0000")
        ws.cell(row=row, column=5).value = tt[4]
        ws.cell(row=row, column=6).value = tt[5]
        row = row + 1
        if row % 10 == 0: # сохраняем 1 раз на 10 строк
            wb.save(wbName)

    if row % 10 > 0: # финальное сохранение
        wb.save(wbName)

# ------------------------------------------------------------ #
def GISgo1_plusplus(rq_number): #проверка запроса: нет/да, люди, файлы
    global GLOBAL_RETRIES_COUNT
    
    try: ## глобальная обработка исключений
        ## ждать загрузки формы поиска
        btn = None
        wait = WebDriverWait(GIS, 30)

        search_btn = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "button[type='submit']")))

        ## запустить поиск запроса по номеру
        NZ_clear_btn = GIS.find_element(By.CSS_SELECTOR, "span.form-base__form-control-clear")
        if "display: block" in NZ_clear_btn.get_attribute("style"):
            NZ_clear_btn.click()
        NZ = GIS.find_element(By.CSS_SELECTOR, "input[type='text'][placeholder='Введите номер запроса']")
        ## имитиация движения мыши
        action = ActionChains(GIS)
        action.move_to_element(NZ)
        action.perform()

        NZ.send_keys(rq_number)
        GIS.execute_script("arguments[0].click();", search_btn)

        ## ждать результата поиска
        results = GIS.find_elements(By.CSS_SELECTOR, "div.section-base__body")[1]
        rq_found = False
        rq_link_href = None
        try:
            wait = WebDriverWait(GIS, 10)
            rq_found = wait.until(EC.text_to_be_present_in_element((By.CSS_SELECTOR, "div.app-content-wrapper"), "Запрос № " + rq_number))
            wait = WebDriverWait(GIS, 30)
        except:
            pass
        ## запрос не найден
        if not rq_found:
            if "Отсутствуют результаты поиска" in results.get_attribute("innerText"):
                return "not_found"
            else:
                return "some_shit"

        ## запрос найден
        ## проверить статус запроса
        icon_debtreq_status = GIS.find_element(By.CSS_SELECTOR, "i.icon-debtreq-status")
        icon_class = icon_debtreq_status.get_attribute("class")
        if "icon-debtreq-status__subrequest_not-sent" in icon_class:
            pass
        elif "icon-debtreq-status__subrequest_sent" in icon_class:
            pass
        elif "icon-debtreq-status__subrequest_generated-automatically" not in icon_class:
            return "sent_auto"
        else:
            return "unknown"

        ##  открыть запрос в новом окне и отработать его
        mainwindow_handle = GIS.current_window_handle
        rq_link_href = GIS.find_element(By.CLASS_NAME, "register-card__header-title").get_attribute("href")
        GIS.execute_script("window.open(arguments[0]);", rq_link_href)
        GIS.switch_to.window(GIS.window_handles[1])

        wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.section-base__footer")))

        section_base_header = GIS.find_element(By.CSS_SELECTOR, "div.section-base__header")

        div_debtreq_status = section_base_header.find_element(By.CSS_SELECTOR, "b.ng-binding")
        debtreq_status = div_debtreq_status.get_attribute("innerText")

        div_debtreq_result_status = section_base_header.find_element(By.CSS_SELECTOR, "div.debtreq__result-type")
        class_debtreq_result_status = div_debtreq_result_status.get_attribute("class")
        if "debtreq__result-type--4" in class_debtreq_result_status:
            debtreq_result_status = "Задолженности нет;"

        elif "debtreq__result-type--3" in class_debtreq_result_status:
            debtreq_result_status = "Есть информация о задолженности"
            act_attached = True
            for person_info in GIS.find_elements(By.CSS_SELECTOR, "debtreq-debt-person-info"):
                fio = person_info.get_attribute("innerText").split("\n")[1]
                act = "Копия судебного акта" in person_info.get_attribute("innerText")
                if not act: act_attached = False
            debtreq_result_status = debtreq_result_status + ";" + ["Файлов нет", "Файлы есть"][act_attached]
            
        elif "debtreq__result-type--2" in class_debtreq_result_status:
            debtreq_result_status = class_debtreq_result_status + ";;"

        elif "debtreq__result-type--1" in class_debtreq_result_status:
            debtreq_result_status = "Ответ не добавлен;"

        else:
            debtreq_result_status = class_debtreq_result_status + ";;"

        section_base_content = GIS.find_elements(By.CSS_SELECTOR, "div.section-base__content")
        sbc = section_base_content[0]
        deadline = sbc.find_element(By.CSS_SELECTOR, "table > tbody > tr > td:nth-child(4)").get_attribute("innerText")
        sbc = section_base_content[1]
        address = sbc.find_element(By.CSS_SELECTOR, "table > tbody > tr:nth-child(4) > td:nth-child(2)").get_attribute("innerText")

        ## закрыть окно
        GIS.close()

        ## вернуться в главное окно
        GIS.switch_to.window(mainwindow_handle)

        GLOBAL_RETRIES_COUNT = 0

        return f"{rq_number}" + ";" + debtreq_status + ";" + debtreq_result_status + ";" + address + ";" + deadline

    except:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"При выполнении возникло исключение {exc_type}\n"
            f"\tописание:\t{exc_obj}\n"
            f"\tмодуль:\t\t{fname}\tстрока:\t{exc_tb.tb_lineno}")

        GLOBAL_RETRIES_COUNT = GLOBAL_RETRIES_COUNT + 1
        if GLOBAL_RETRIES_COUNT > 3: return "FATAL"
        else: return "ERROR"
        
        
RUN_FROM_IDLE = "idlelib" in sys.modules
main()
