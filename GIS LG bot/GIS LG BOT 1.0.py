import os, re, sys, time
import inspect

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
##from selenium.webdriver.common.by import By
##from selenium.webdriver.support import expected_conditions as EC
##from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains

GIS = None
sessionid = ""
RunFromIDLE = None
  
##------------------------------------------------------------##
def GISinit():
    global GIS, sessionid, RunFromIDLE
    chrome_options = Options()
    if not RunFromIDLE: 
        chrome_options.add_argument("--headless")
    chrome_options.add_argument("user-data-dir=" + os.getenv("TEMP") + "\\gis1")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    GIS = webdriver.Chrome(options = chrome_options)
    GIS.set_page_load_timeout(180)
    print("Начало работы с сайтом")
    GIS.get("https://dom.gosuslugi.ru/404")
    GIS.add_cookie({'name' : 'sessionid', 'value' : sessionid, 'domain' : '.dom.gosuslugi.ru'})
    print("Загрузка первой страницы (может быть довольного долго)")
    GIS.get("https://my.dom.gosuslugi.ru/organization-cabinet/#!/debts/received-requests")
    
def log(s):
    print(s)
    with open('log.txt', 'a') as f: f.write(f'{s}\n')

##########################################################################
try:
    f = open("sessionid.txt", "r")
    sessionid = f.readline()[:34]
    f.close()
except:
    print("Проблема с чтением идентификатора сессии из файла")
    time.sleep(14400)
    raise SystemExit

RunFromIDLE = "idlelib" in sys.modules
GISinit()

t00 = time.monotonic()
print("Сайт загружен, начало работы со списком запросов: ", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
wbName = 'lg_no.xlsx'
wb = load_workbook(filename = wbName)
ws1 = wb.worksheets[0]
r = 2 # строка в исходном списке
rmax = ws1.max_row - 1 # всего строк
cc = ws1.cell(row = r , column = 1).value
status = 0
while cc != None:
    t0 = time.monotonic()
    if status == 999:
        log(f"Запрос № {cc} - проблема при обработке запроса")
        status = 0
    ## ожидание загрузки формы поиска
    status = 0; wait_count = 0; btn = None
    while True:
        if wait_count > 10: status = 999; break
        try: 
            btn = GIS.find_element_by_css_selector("button[type='submit']"); break
        except:
            time.sleep(1); wait_count += 1
    if status == 999: print( "Break at line no. ", inspect.getframeinfo(inspect.currentframe()).lineno); break
    NZ = GIS.find_element_by_css_selector("input[type='text'][placeholder='Введите номер запроса']")
    NZ.clear()
    NZ.send_keys(cc)
    GIS.execute_script("arguments[0].click();", btn)

    ## ожидание результата поиска
    status = 0; wait_count = 0; rq_link = None
    while True:
        if wait_count > 10: status = 999; break
        try:
            rq_link = GIS.find_element_by_xpath(".//span[contains(text(),'Запрос № " + cc +"')]"); break
        except:
            time.sleep(1); wait_count += 1
    if status == 999: ## запрос по номеру не найден или ошибка
        ## проверим "Отсутствуют результаты поиска"
        status = 0; wait_count = 0; norq = None
        while True:
            if wait_count > 3: status = 999; break
            try:
                norq = GIS.find_element_by_xpath(".//div[contains(text(),'Отсутствуют результаты поиска')]"); break
            except:
                time.sleep(1); wait_count += 1
    if status == 999: print( "Break at line no. ", inspect.getframeinfo(inspect.currentframe()).lineno); break

    if rq_link != None:
        if not RunFromIDLE: action = ActionChains(GIS); action.move_to_element(rq_link).perform()
    ##  проверить статус запроса - не был отправлен ранее
        status = 0; wait_count = 0; icon_debtreq_status = None
        while True:
            if wait_count > 10: status = 999; break
            try:
                icon_debtreq_status = GIS.find_element_by_class_name("icon-debtreq-status"); break
            except:
                time.sleep(1); wait_count += 1
        if status == 999: print( "Break at line no. ", inspect.getframeinfo(inspect.currentframe()).lineno); break
        if icon_debtreq_status == None: break
        if "icon-debtreq-status__subrequest_not-sent" not in icon_debtreq_status.get_attribute("class"):
            log(f"Запрос № {cc} ({r - 1} из {rmax}) уже был отработан ранее, переходим к следующему. Потеряли {round(time.monotonic() - t0, 2)} сек.")
            r = r + 1
            cc = ws1.cell(row = r, column = 1).value
            continue

    ##  Открыть запрос в новом окне и отработать его
        mainwindow_handle = GIS.current_window_handle
        GIS.execute_script("arguments[0].click();", rq_link)
        GIS.switch_to.window(GIS.window_handles[1])
        status = 0; wait_count = 0; btn = None
        while True:
            if wait_count > 10: status = 999; break
            try:
                btn = GIS.find_element_by_class_name("btn-action"); break
            except:
                time.sleep(1); wait_count += 1
        if status == 999: print( "Break at line no. ", inspect.getframeinfo(inspect.currentframe()).lineno); break
        GIS.execute_script("arguments[0].click();", btn)

    ##ждём диалоговое окно "Добавление ответа на запрос № ********* от **.**.****"с выбором ответа да нет для указания наличия задолженности
        status = 0; wait_count = 0; modal_dialog = None
        while True:
            if wait_count > 15: status = 999; break
            try:
                modal_dialog = GIS.find_element_by_class_name("modal-dialog"); break
            except:
                time.sleep(1); wait_count += 1
        if status == 999: print( "Break at line no. ", inspect.getframeinfo(inspect.currentframe()).lineno); break
            
    ##нажимаем нет - задолженности нет
        status = 0; wait_count = 0; radio = None
        while True:
            if wait_count > 10: status = 999; break
            try:
                radio = GIS.find_elements_by_xpath(".//input[@type='radio']")[1]; break
            except:
                time.sleep(1); wait_count += 1
        if status == 999: print( "Break at line no. ", inspect.getframeinfo(inspect.currentframe()).lineno); break
        GIS.execute_script("arguments[0].click();", radio)    

    ##нажимаем Сохранить ответ
        status = 0; wait_count = 0; btn = None
        while True:
            if wait_count > 10: status = 999; break
            try:
                btn = GIS.find_element_by_xpath(".//button[contains(text(),'Сохранить ответ')]"); break
            except:
                time.sleep(1); wait_count += 1
        if status == 999: print( "Break at line no. ", inspect.getframeinfo(inspect.currentframe()).lineno); break
        GIS.execute_script("arguments[0].click();", btn)    
       
    ##ждём закрытия диалога
        status = 0; wait_count = 0
        while True:
            if wait_count > 15: status = 999; break
            try:
                temp = modal_dialog.get_attribute("innerText")
                time.sleep(1); wait_count += 1
            except:
                break
        if status == 999: print( "Break at line no. ", inspect.getframeinfo(inspect.currentframe()).lineno); break

    ##нажимаем отправить ответ
        status = 0; wait_count = 0; btn = None
        while True:
            if wait_count > 10: status = 999; break
            try:
                btn = GIS.find_element_by_xpath(".//button[contains(text(),'Отправить ответ')]"); break
            except:
                time.sleep(1); wait_count += 1
        if status == 999: print( "Break at line no. ", inspect.getframeinfo(inspect.currentframe()).lineno); break
        GIS.execute_script("arguments[0].click();", btn)    

    ##ждём появления диалога "Подтверждение"
        status = 0; wait_count = 0; modal_dialog = None
        while True:
            if wait_count > 10: status = 999; break
            try:
                modal_dialog = GIS.find_element_by_class_name("modal-dialog"); break
            except:
                time.sleep(1); wait_count += 1
        if status == 999: print( "Break at line no. ", inspect.getframeinfo(inspect.currentframe()).lineno); break
        
    ##нажимаем Да
        status = 0; wait_count = 0; btn = None
        while True:
            if wait_count > 10: status = 999; break
            try:
                btn = GIS.find_element_by_xpath(".//button/span[contains(text(),'Да')]"); break
            except:
                time.sleep(1); wait_count += 1
        if status == 999: print( "Break at line no. ", inspect.getframeinfo(inspect.currentframe()).lineno); break
        GIS.execute_script("arguments[0].click();", btn)    
        
    ##ждём закрытия диалога
        status = 0; wait_count = 0
        while True:
            if wait_count > 10: status = 999; break
            try:
                temp = modal_dialog.get_attribute("innerText")
                time.sleep(1); wait_count += 1
            except:
                break
        if status == 999: print( "Break at line no. ", inspect.getframeinfo(inspect.currentframe()).lineno); break
            
    ##ждём появления диалога "Операция выполнена успешно"
        status = 0; wait_count = 0; modal_dialog = None
        while True:
            if wait_count > 10: status = 999; break
            try:
                modal_dialog = GIS.find_element_by_class_name("modal-dialog"); break
            except:
                time.sleep(1); wait_count += 1
        if status == 999: print( "Break at line no. ", inspect.getframeinfo(inspect.currentframe()).lineno); break
        if "Операция выполнена успешно" not in modal_dialog.get_attribute("innerText"): break
        
    ##просто закрываем окно
        GIS.close(); GIS.switch_to.window(mainwindow_handle) # вернуться в главное окно
        
        log(f"Запрос № {cc} ({r - 1} из {rmax}) успешно отработан за {round(time.monotonic() - t0, 2)} сек.")
    elif norq != None:
        log(f"Запрос № {cc} ({r - 1} из {rmax}) не найден (отозван или сформирован другим пользователем, но не отправлен), переходим к следующему. Потеряли {round(time.monotonic() - t0, 2)} сек.")
    r = r + 1
    cc = ws1.cell(row = r, column = 1).value

print("Обработка завершена: ", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
t = round(time.monotonic() - t00, 0)
try: print(f"С момента запуска прошло {timedelta(seconds = t)}, в среднем {round(t / (r - 2), 2)} сек. на запрос")
except: pass
